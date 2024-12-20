import os
from flask import Blueprint, request, render_template, redirect, url_for, session
from openpyxl.reader.excel import load_workbook

from app.models import cargar_usuarios, guardar_usuario, cargar_historial_usuario, guardar_datos_usuario
from app.utils.utils import model

usuario = Blueprint('main', __name__)


@usuario.route('/register', methods=['GET', 'POST'])
def register():
    error = None
    success = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        usuarios = cargar_usuarios()

        if username in usuarios:
            error = 'El nombre de usuario ya existe.'
        else:
            guardar_usuario(username, password)
            success = 'Usuario registrado con éxito. ¡Ahora puedes iniciar sesión!'
            return redirect(url_for('main.login'))

    return render_template('register.html', error=error, success=success)


@usuario.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        usuarios = cargar_usuarios()

        if username in usuarios and usuarios[username] == password:
            session['logged_in'] = True
            session['username'] = username
            session['user_data'] = cargar_historial_usuario(username)
            return redirect(url_for('chat_ia.chat'))
        else:
            error = 'Usuario o contraseña incorrectos.'

    return render_template('login.html', error=error)


""""@usuario.route('/', methods=['GET', 'POST'])
@usuario.route('/chat', methods=['GET', 'POST'])
def chat():
    if not session.get('logged_in'):
        return redirect(url_for('main.login'))

    respuesta = ""
    if request.method == 'POST':
        pregunta = request.form['pregunta']
        historial = session.get('user_data', [])
        historial_texto = " ".join([f"Pregunta: {h['Pregunta']} Respuesta: {h['Respuesta']}" for h in historial])
        contexto = historial_texto + " " + pregunta
        respuesta = model.generate_content(contexto).text

        historial.append({'Pregunta': pregunta, 'Respuesta': respuesta})
        session['user_data'] = historial
        guardar_datos_usuario(session['username'], pregunta, respuesta)

    return render_template('form.html', salida=respuesta)"""


@usuario.route('/ver_problemas')
def ver_problemas():
    try:
        file_path = os.path.join(os.path.dirname(__file__), 'arbol_problemas.xlsx')
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]

        # Crear la tabla HTML manualmente respetando celdas unidas
        html_table = "<table class='table table-bordered'>"

        for row in ws.iter_rows():
            html_table += "<tr>"
            for cell in row:
                # Detectar si la celda está en una celda combinada
                merged_cell = None
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        merged_cell = merged_range
                        break

                if merged_cell and cell.coordinate == merged_cell.coord.split(":")[0]:
                    # Si la celda es el inicio de una celda unida
                    min_col, min_row, max_col, max_row = merged_cell.bounds
                    rowspan = max_row - min_row + 1
                    colspan = max_col - min_col + 1
                    html_table += f"<td rowspan='{rowspan}' colspan='{colspan}'>{cell.value}</td>"
                elif merged_cell:
                    # Si la celda está dentro de un área unida, no mostrar nada
                    continue
                else:
                    # Si es una celda normal
                    html_table += f"<td>{cell.value}</td>"

            html_table += "</tr>"

        html_table += "</table>"

    except Exception as e:
        html_table = f"Error al cargar el archivo: {str(e)}"

    return render_template('ver_problemas.html', tabla_excel=html_table)


@usuario.route('/ver_objetivos')
def ver_objetivos():
    try:
        # Cargar el archivo Excel usando openpyxl
        file_path = os.path.join(os.path.dirname(__file__), 'arbol_objetivos.xlsx')
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]

        # Crear la tabla HTML manualmente respetando celdas unidas
        html_table = "<table class='table table-bordered'>"

        for row in ws.iter_rows():
            html_table += "<tr>"
            for cell in row:
                # Detectar si la celda está en una celda combinada
                merged_cell = None
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        merged_cell = merged_range
                        break

                if merged_cell and cell.coordinate == merged_cell.coord.split(":")[0]:
                    # Si la celda es el inicio de una celda unida
                    min_col, min_row, max_col, max_row = merged_cell.bounds
                    rowspan = max_row - min_row + 1
                    colspan = max_col - min_col + 1
                    html_table += f"<td rowspan='{rowspan}' colspan='{colspan}'>{cell.value}</td>"
                elif merged_cell:
                    # Si la celda está dentro de un área unida, no mostrar nada
                    continue
                else:
                    # Si es una celda normal
                    html_table += f"<td>{cell.value}</td>"

            html_table += "</tr>"

        html_table += "</table>"

    except Exception as e:
        html_table = f"Error al cargar el archivo: {str(e)}"

    return render_template('ver_objetivos.html', tabla_excel=html_table)


@usuario.route('/arbolproblema')
def arbolproblema():
    causa_directa = session.get('causas_directas', [])
    causa_indirecta = session.get('causas_indirectas', [])
    efectos_indirectos = session.get('efectos_indirectos', [])
    efectos_directos = session.get('efectos_directos', [])

    return render_template('arbol_problema.html',
                           causas_directas=causa_directa,
                           causas_indirectas=causa_indirecta,
                           efectos_directos=efectos_indirectos,
                           efectos_indirectos=efectos_directos)


@usuario.route('/arbolobjetivos')
def arbolobjetivos():
    fines_directos = session.get('fines_directos', [])
    fines_indirectos = session.get('fines_indirectos', [])
    objetivos_especificos = session.get('objetivos_especificos', [])
    medios = session.get('medios', [])

    return render_template('arbol_objetivos.html', problema="problema", fines_directos=fines_directos,
                           fines_indirectos=fines_indirectos, objetivos_especificos=objetivos_especificos,
                           medios=medios)


@usuario.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('main.login'))
